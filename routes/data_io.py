"""
Fitur Import & Export (spesifikasi bagian 9 & 10).

IMPORT:
    Membaca file .xlsx yang strukturnya mengikuti template spreadsheet asli
    (sheet 'Biaya Tetap', 'Biaya Variabel', 'Biaya Tidak Langsung', 'Inventori Harian').
    Ini bukan pengganti total mesin hitung Python (rumus tetap dihitung oleh
    services/*), melainkan cara cepat memasukkan/menyinkronkan data massal
    ketika pemilik UMKM mengubah datanya di Excel.

EXPORT:
    - Excel: menuliskan ULANG data + rumus (bukan angka hasil hardcode) supaya
      file tetap bisa dihitung ulang di Excel, sesuai praktik terbaik.
    - PDF: laporan ringkas hasil perhitungan (read-only, untuk dicetak/dibagikan).
"""

import os
from datetime import datetime

from flask import Blueprint, render_template, redirect, url_for, flash, send_file, current_app
from openpyxl import load_workbook, Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from models import db, FixedCost, OverheadCost, Product, VariableCostComponent, Material, CashFlowEntry
from forms import UploadExcelForm
from services.cost_service import total_fixed_cost, total_overhead_cost
from services.inventory_service import calculate_all_materials
from services.finance_service import calculate_all_cashflow, cashflow_totals
from utils.validators import allowed_file

bp = Blueprint("data_io", __name__)


@bp.route("/import", methods=["GET", "POST"])
def import_excel():
    form = UploadExcelForm()
    if form.validate_on_submit():
        f = form.file.data
        if not allowed_file(f.filename, current_app.config["ALLOWED_UPLOAD_EXTENSIONS"]):
            flash("Format file tidak didukung. Hanya .xlsx.", "danger")
            return redirect(url_for("data_io.import_excel"))

        filepath = os.path.join(current_app.config["UPLOAD_FOLDER"], f.filename)
        f.save(filepath)

        try:
            summary = _import_workbook(filepath)
            flash(
                "Import berhasil: " + ", ".join(f"{k}: {v}" for k, v in summary.items()),
                "success",
            )
        except Exception as exc:  # noqa: BLE001 - tampilkan pesan ke user, jangan crash
            flash(f"Gagal memproses file: {exc}", "danger")

        return redirect(url_for("data_io.import_excel"))

    return render_template("data_io.html", form=form)


def _import_workbook(filepath):
    """
    Membaca sheet-sheet standar dan melakukan UPSERT sederhana ke database.
    Baris dikenali lewat nama kategori/produk (kolom pertama/kedua yang relevan),
    bukan lewat posisi baris, supaya tahan terhadap penyisipan baris baru.
    """
    wb = load_workbook(filepath, data_only=True)
    summary = {}

    if "Biaya Tetap" in wb.sheetnames:
        ws = wb["Biaya Tetap"]
        count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            category, amount = row[1], row[2]
            if not category or category == "TOTAL FIXED COST":
                continue
            existing = FixedCost.query.filter_by(category=str(category)).first()
            if existing:
                existing.amount_per_month = float(amount or 0)
            else:
                db.session.add(FixedCost(category=str(category), amount_per_month=float(amount or 0)))
            count += 1
        summary["Biaya Tetap"] = count

    if "Biaya Tidak Langsung" in wb.sheetnames or "Biaya Tidak Langsung " in wb.sheetnames:
        name = "Biaya Tidak Langsung" if "Biaya Tidak Langsung" in wb.sheetnames else "Biaya Tidak Langsung "
        ws = wb[name]
        count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            category, amount, pct = row[1], row[2], row[3]
            if not category or "TOTAL" in str(category).upper():
                continue
            existing = OverheadCost.query.filter_by(category=str(category)).first()
            if existing:
                existing.amount_per_month = float(amount or 0)
                existing.allocation_pct = float(pct or 0)
            else:
                db.session.add(
                    OverheadCost(category=str(category), amount_per_month=float(amount or 0), allocation_pct=float(pct or 0))
                )
            count += 1
        summary["Biaya Tidak Langsung"] = count

    if "Biaya Variabel" in wb.sheetnames:
        ws = wb["Biaya Variabel"]
        count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            product_name, component_name, unit, qty, price = row[1], row[2], row[3], row[4], row[5]
            if not product_name or not component_name:
                continue
            product = Product.query.filter_by(name=str(product_name)).first()
            if not product:
                product = Product(name=str(product_name), selling_price=0)
                db.session.add(product)
                db.session.flush()
            existing = VariableCostComponent.query.filter_by(
                product_id=product.id, component_name=str(component_name)
            ).first()
            if existing:
                existing.unit = str(unit or "")
                existing.qty_per_unit = float(qty or 0)
                existing.price_per_unit = float(price or 0)
            else:
                db.session.add(
                    VariableCostComponent(
                        product_id=product.id,
                        component_name=str(component_name),
                        unit=str(unit or ""),
                        qty_per_unit=float(qty or 0),
                        price_per_unit=float(price or 0),
                    )
                )
            count += 1
        summary["Biaya Variabel (komponen)"] = count

    if "Inventori Harian" in wb.sheetnames:
        ws = wb["Inventori Harian"]
        count = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            # B..K: nama, satuan, avg demand, max demand, avg LT, max LT, order cost, holding cost, price, stock
            name = row[1]
            if not name:
                continue
            existing = Material.query.filter_by(name=str(name)).first()
            values = dict(
                unit=str(row[2] or ""),
                avg_demand_month=float(row[3] or 0),
                max_demand_month=float(row[4] or 0),
                avg_lead_time_days=float(row[5] or 0),
                max_lead_time_days=float(row[6] or 0),
                order_cost=float(row[7] or 0),
                holding_cost_per_unit_month=float(row[8] or 0),
                purchase_price=float(row[9] or 0),
                current_stock=float(row[10] or 0),
            )
            if existing:
                for k, v in values.items():
                    setattr(existing, k, v)
            else:
                db.session.add(Material(name=str(name), **values))
            count += 1
        summary["Inventori Harian (bahan baku)"] = count

    db.session.commit()
    return summary


@bp.route("/export/excel")
def export_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Biaya Tetap"
    ws.append(["No", "Kategori", "Nominal per Bulan (Rp)"])
    for i, item in enumerate(FixedCost.query.all(), start=1):
        ws.append([i, item.category, item.amount_per_month])
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=2, value="TOTAL")
    ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row - 1})")

    ws2 = wb.create_sheet("Biaya Tidak Langsung")
    ws2.append(["No", "Kategori", "Nominal per Bulan (Rp)", "% Alokasi", "Overhead Teralokasi (Rp)"])
    for i, item in enumerate(OverheadCost.query.all(), start=1):
        row = i + 1
        ws2.append([i, item.category, item.amount_per_month, item.allocation_pct, f"=C{row}*D{row}"])
    total_row2 = ws2.max_row + 1
    ws2.cell(row=total_row2, column=2, value="TOTAL")
    ws2.cell(row=total_row2, column=5, value=f"=SUM(E2:E{total_row2 - 1})")

    ws3 = wb.create_sheet("Biaya Variabel")
    ws3.append(["Produk", "Komponen", "Satuan", "Qty/Unit", "Harga/Satuan (Rp)", "Biaya/Unit (Rp)"])
    for c in VariableCostComponent.query.join(Product).all():
        row = ws3.max_row + 1
        ws3.append([c.product.name, c.component_name, c.unit, c.qty_per_unit, c.price_per_unit, f"=D{row}*E{row}"])

    ws4 = wb.create_sheet("Inventori")
    ws4.append([
        "Bahan Baku", "Satuan", "Permintaan Harian Rata2", "Permintaan Harian Maks",
        "Permintaan Tahunan", "EOQ", "Safety Stock", "Reorder Point", "Status EOQ",
        "Kebutuhan Bulan Depan", "SS LFL", "Jumlah Pesan LFL", "Status LFL",
    ])
    for r in calculate_all_materials(Material.query.all()):
        ws4.append([
            r["material_name"], r["unit"], r["permintaan_harian_rata2"], r["permintaan_harian_maksimum"],
            r["permintaan_tahunan"], r["eoq"], r["safety_stock"], r["reorder_point"], r["status_reorder_eoq"],
            r["kebutuhan_bulan_depan_lfl"], r["safety_stock_lfl"], r["jumlah_pesan_lfl"], r["status_reorder_lfl"],
        ])

    ws5 = wb.create_sheet("Arus Kas")
    ws5.append([
        "Bulan", "Produk", "Unit Terjual", "Harga Jual", "Revenue", "Var Cost/Unit",
        "Total Var Cost", "Alokasi Fixed", "Alokasi Overhead", "Total Cost", "Profit", "Margin (%)",
    ])
    for r in calculate_all_cashflow(CashFlowEntry.query.all()):
        ws5.append([
            r["month_label"], r["product_name"], r["units_sold"], r["selling_price"], r["revenue"],
            r["variable_cost_per_unit"], r["total_variable_cost"], r["allocated_fixed_cost"],
            r["allocated_overhead"], r["total_cost"], r["profit"], r["profit_margin_pct"],
        ])

    filename = f"export_umkm_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(current_app.config["EXPORT_FOLDER"], filename)
    wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name=filename)


@bp.route("/export/pdf")
def export_pdf():
    filename = f"laporan_umkm_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(current_app.config["EXPORT_FOLDER"], filename)

    doc = SimpleDocTemplate(filepath, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph("Laporan Inventori & Finansial UMKM", styles["Title"]), Spacer(1, 12)]

    elements.append(Paragraph(f"Total Biaya Tetap: Rp {total_fixed_cost():,.0f}".replace(",", "."), styles["Normal"]))
    elements.append(Paragraph(f"Total Overhead Teralokasi: Rp {total_overhead_cost():,.0f}".replace(",", "."), styles["Normal"]))

    cf_totals = cashflow_totals(CashFlowEntry.query.all())
    elements.append(Paragraph(f"Total Revenue: Rp {cf_totals['total_revenue']:,.0f}".replace(",", "."), styles["Normal"]))
    elements.append(Paragraph(f"Total Profit: Rp {cf_totals['total_profit']:,.0f}".replace(",", "."), styles["Normal"]))
    elements.append(Spacer(1, 16))

    elements.append(Paragraph("Status Inventori Bahan Baku", styles["Heading2"]))
    data = [["Bahan Baku", "EOQ", "Safety Stock", "ROP", "Status EOQ", "Status LFL"]]
    for r in calculate_all_materials(Material.query.all()):
        data.append([r["material_name"], r["eoq"], r["safety_stock"], r["reorder_point"], r["status_reorder_eoq"], r["status_reorder_lfl"]])
    table = Table(data, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f6f5c")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    return send_file(filepath, as_attachment=True, download_name=filename)
